import openpyxl

file_path = 'CHỐT FILE UP LUONG 03.2026 (06.04).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
        print(" | ".join([str(cell) if cell is not None else "" for cell in row]))
